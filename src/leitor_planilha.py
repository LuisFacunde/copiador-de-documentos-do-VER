from pathlib import Path
import openpyxl


def ler_prontuarios(
    caminho: Path,
    aba: str | None = None,
    coluna: str = 'PRONT',
) -> list[str]:
    caminho = Path(caminho)
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {caminho}")

    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[aba] if aba and aba in wb.sheetnames else wb.active

    cabecalho = None
    for row in ws.iter_rows(values_only=True):
        if any(v is not None for v in row):
            cabecalho = [str(v).strip() if v is not None else '' for v in row]
            break

    if cabecalho is None or coluna not in cabecalho:
        wb.close()
        raise ValueError(
            f"Coluna '{coluna}' não encontrada na planilha. "
            f"Colunas disponíveis: {cabecalho}"
        )

    indice_coluna = cabecalho.index(coluna)

    prontuarios = []
    cabecalho_encontrado = False

    for row in ws.iter_rows(values_only=True):
        if not cabecalho_encontrado:
            valores = [str(v).strip() if v is not None else '' for v in row]
            if coluna in valores:
                cabecalho_encontrado = True
            continue

        valor = row[indice_coluna]
        if valor is not None:
            prontuario = str(valor).strip().split('.')[0]
            if prontuario:
                prontuarios.append(prontuario)

    wb.close()

    vistos = set()
    unicos = []
    for p in prontuarios:
        if p not in vistos:
            vistos.add(p)
            unicos.append(p)

    return unicos
